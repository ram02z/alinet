from dataclasses import dataclass, field
from transformers import HfArgumentParser, set_seed
import openpyxl
from openpyxl import Workbook
import json
import os
import sys

SRC_DIR = os.path.join(os.path.dirname(__file__), "src")
sys.path.append(SRC_DIR)
from alinet import chunking, qg, rag  # noqa: E402


def create_or_load_workbook(filename):
    try:
        workbook = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        workbook = Workbook()
    return workbook


@dataclass
class EvaluateRAGArguments:
    lecture_path: str = field(
        metadata={"help": "Name of lecture to store data in excel worksheet"}
    )
    top_k: int = field(
        default=1,
        metadata={"help": "Number of relevant contexts to retrieve"},
    )
    distance_threshold: float = field(
        default=0,
        metadata={"help": "Distance threshold to consider a context as relevant"},
    )
    collection_name: str = field(
        default="default",
        metadata={"help": "The name of the collection in the vectordb"},
    )
    seed: int = field(default=42, metadata={"help": "Random seed"})


def main():
    parser = HfArgumentParser((EvaluateRAGArguments))
    args = parser.parse_args_into_dataclasses()[0]
    print("Lecture Path: ", args.lecture_path)
    # list to store files
    doc_paths = []
    transcript_path = ""
    # Iterate directory
    for path in os.listdir(args.lecture_path):
        full_path = os.path.join(args.lecture_path, path)

        if os.path.isfile(full_path):
            if path.endswith(".pdf"):
                doc_paths.append(full_path)
            # check if the file is a JSON
            elif path.endswith(".json"):
                transcript_path = full_path

    print("Supplementary Docs: ", doc_paths)
    print("Json Path: ", transcript_path)

    set_seed(args.seed)
    db = rag.Database()

    collection = db.create_collection()

    # Store documents in collection
    pdfs_bytes: list[bytes] = []
    for doc_path in doc_paths:
        with open(doc_path, "rb") as f:
            pdf_bytes = f.read()
        pdfs_bytes.append(pdf_bytes)

    db.store_documents(collection, pdfs_bytes=pdfs_bytes)

    # Path to transcript
    with open(transcript_path, "rb") as file:
        data = json.load(file)
        whisper_chunks = data["chunks"]
        duration = data["duration"]

    chunk_pipe = chunking.Pipeline(qg.Model.BALANCED_RA)
    transcript_chunks = chunk_pipe(whisper_chunks, duration, stride_length=32)

    source_texts = [chunk.text for chunk in transcript_chunks]

    sources_with_context = db.add_relevant_context_to_sources(
        source_texts=source_texts,
        collection=collection,
        top_k=args.top_k,
        distance_threshold=args.distance_threshold,
    )

    qg_pipe = qg.Pipeline(qg.Model.BALANCED_RA)

    # Run QG Model On Source Text
    source_questions = qg_pipe(source_texts)

    # Run QG Model On Source + RAG Text
    source_rag_questions = qg_pipe(sources_with_context)

    rag_texts = []
    # Extract the rag text from the combined texts
    for i, text_with_context in enumerate(sources_with_context):
        original_text = source_texts[i]
        rag_only_text = text_with_context.split(original_text, 1)[-1].strip()
        rag_texts.append(rag_only_text)

    # Run QG Model On RAG Text
    rag_questions = qg_pipe(rag_texts)

    text_data = []
    for i, text_with_context in enumerate(sources_with_context):
        source_text = source_texts[i]
        source_question = source_questions[i]

        rag_text = rag_texts[i]
        rag_question = rag_questions[i]

        source_rag_question = source_rag_questions[i]
        text_data.append(
            (
                source_text,
                source_question,
                "",
                rag_text,
                rag_question,
                "",
                text_with_context,
                source_rag_question,
                "",
            )
        )

    # Create or load the workbook
    workbook = create_or_load_workbook("source_vs_rag.xlsx")

    # Create a new worksheet with the lecture name
    _, file = os.path.split(args.lecture_path)
    print("Worksheet: ", file)
    worksheet = workbook.create_sheet(file)

    # Write the header row
    worksheet["A1"] = "Source"
    worksheet["B1"] = "Source Question"
    worksheet["C1"] = "Quality"
    worksheet["D1"] = "RAG"
    worksheet["E1"] = "RAG Question"
    worksheet["F1"] = "Quality"
    worksheet["G1"] = "Source + RAG"
    worksheet["H1"] = "Source + RAG Question"
    worksheet["I1"] = "Quality"

    # Write the data rows
    for row_num, data in enumerate(text_data, start=2):
        worksheet.cell(row=row_num, column=1, value=data[0])
        worksheet.cell(row=row_num, column=2, value=data[1])
        worksheet.cell(row=row_num, column=3, value=data[2])
        worksheet.cell(row=row_num, column=4, value=data[3])
        worksheet.cell(row=row_num, column=5, value=data[4])
        worksheet.cell(row=row_num, column=6, value=data[5])
        worksheet.cell(row=row_num, column=7, value=data[6])
        worksheet.cell(row=row_num, column=8, value=data[7])

    # Save the workbook
    workbook.save("source_vs_rag.xlsx")

    # Delete database
    db.client.reset()


if __name__ == "__main__":
    main()
