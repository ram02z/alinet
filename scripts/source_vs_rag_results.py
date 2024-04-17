# Read through the Excel file
# File_path: sample_data/source_vs_rag.xlsx

# The Excel file contains 10 worksheets each specific to a lecture.

# A2 to O2 are the headings of each Column

# Source Text, Source Question
# RAG Text, RAG Question
# Source Text, Source + RAG Question

# Each question for each text has a quality rating from 3 different rates.
# The quality is judged using 4 comma separated values each representing a different metric.
# A question is considered BAD Quality if there is at-least one score less than 3 or if all values are 3. If this isn't met then its a GOOD Quality question.
#   - For example, different question ratings:
#   - 3,3,3,2 - Negative: Have at-least one value that is less than 3.
#   - 3,2,3,3 - Negative
#   - 3,2,3,5 - Negative: Have at-least one value less than 3, even though another value is super positive we don't care since it failed one criteria.
#   - 3,3,3,3 - Negative: If all neutral, it probably means uncertainty in labellers decision.
#   - 3,3,3,4 - Positive: All the other values are 3, but we have a 4 so we are leaning towards Positive.
#   - 3,4,3,3 - Positive


# For each Rater, Wil Quality, Omar Quality, and Jia Wei Quality. Calculate using the criteria above whether the question GOOD or BAD Quality.

import openpyxl
from openpyxl.styles import Alignment


SOURCE_INDEX = 2
RAG_INDEX = 8
SOURCE_RAG_INDEX = 14


def calculate_quality(scores_str: str):
    # Incase of typos in the document:
    # Remove spaces in the string and leading and trailing commas.
    clean_score = scores_str.replace(" ", "").strip(",").split(",")

    scores = [int(score) for score in clean_score]
    if any(score < 3 for score in scores) or all(score == 3 for score in scores):
        return "BAD"
    else:
        return "GOOD"


def mode(values: list):
    return max(set(values), key=values.count)


good_source_count = 0
bad_source_count = 0

good_rag_count = 0
bad_rag_count = 0

good_source_rag_count = 0
bad_source_rag_count = 0


# Col 1 is the Source text ------------- Index 0
# Col 2 is the Source question
# Col 3,4,5 are the individual quality  ------------- Index 2
# Col 6 is Source overall quality
#
# Col 7 is the RAG text  ------------- Index 6
# Col 8 is the RAG question
# Col 9,10,11 are the individual quality  ------------- Index 8
# Col 12 is RAG overall quality
#
# Col 13 is the Source + RAG text  ------------- Index 12
# Col 14 is the Source + RAG question
# Col 15,16,17 are the individual quality  ------------- Index 14
# Col 18 is Source + RAG overall quality
def get_quality(index):
    results = [""] * 3
    for i in range(3):
        quality = calculate_quality(row[index + i])
        results[i] = quality
        cell = worksheet.cell(row=row_num, column=index + 1 + i, value=quality)
        cell.alignment = Alignment(vertical="top", horizontal="left")

    mode_value = mode(results)
    cell = worksheet.cell(row=row_num, column=index + 4, value=mode_value)
    cell.alignment = Alignment(vertical="top", horizontal="left")

    return mode_value


good_source_good_rag_count = 0
good_source_bad_rag_count = 0
bad_source_good_rag_count = 0
bad_source_bad_rag_count = 0


source_rag_combinations = {
    "Good Source, Good RAG": {"Good Source + RAG": 0, "Bad Source + RAG": 0},
    "Good Source, Bad RAG": {"Good Source + RAG": 0, "Bad Source + RAG": 0},
    "Bad Source, Good RAG": {"Good Source + RAG": 0, "Bad Source + RAG": 0},
    "Bad Source, Bad RAG": {"Good Source + RAG": 0, "Bad Source + RAG": 0},
}
# Load the Excel file
workbook = openpyxl.load_workbook("sample_data/source_vs_rag.xlsx")

# Loop through each worksheet
for worksheet_name in workbook.sheetnames:
    print("WORKSHEET: ", worksheet_name)
    worksheet = workbook[worksheet_name]

    # Insert new columns for the overall quality results for each text
    worksheet.insert_cols(16)
    cell = worksheet.cell(row=2, column=16, value="Overall Quality")
    cell.alignment = Alignment(vertical="top", horizontal="left")

    worksheet.insert_cols(11)
    cell = worksheet.cell(row=2, column=11, value="Overall Quality")
    cell.alignment = Alignment(vertical="top", horizontal="left")

    worksheet.insert_cols(6)
    cell = worksheet.cell(row=2, column=6, value="Overall Quality")
    cell.alignment = Alignment(vertical="top", horizontal="left")

    # Skip the first 2 rows - Enumerate behaviour causes index to start from 0, so specify it starts from row 3
    for row_num, row in enumerate(
        worksheet.iter_rows(min_row=3, max_row=7, values_only=True), start=3
    ):
        # Source starts from Index 2 (Column 3)
        source_quality = get_quality(SOURCE_INDEX)
        # RAG starts from index 8 (Column 9)
        rag_quality = get_quality(RAG_INDEX)
        # Source + RAG starts from Index 14 (Column 15)
        source_rag_quality = get_quality(SOURCE_RAG_INDEX)

        if source_quality == "GOOD" and rag_quality == "GOOD":
            good_source_good_rag_count += 1

            if source_rag_quality == "GOOD":
                source_rag_combinations["Good Source, Good RAG"][
                    "Good Source + RAG"
                ] += 1

            elif source_rag_quality == "BAD":
                source_rag_combinations["Good Source, Good RAG"][
                    "Bad Source + RAG"
                ] += 1

        elif source_quality == "GOOD" and rag_quality == "BAD":
            good_source_bad_rag_count += 1

            if source_rag_quality == "GOOD":
                source_rag_combinations["Good Source, Bad RAG"][
                    "Good Source + RAG"
                ] += 1

            elif source_rag_quality == "BAD":
                source_rag_combinations["Good Source, Bad RAG"]["Bad Source + RAG"] += 1

        elif source_quality == "BAD" and rag_quality == "GOOD":
            bad_source_good_rag_count += 1

            if source_rag_quality == "GOOD":
                source_rag_combinations["Bad Source, Good RAG"][
                    "Good Source + RAG"
                ] += 1

            elif source_rag_quality == "BAD":
                source_rag_combinations["Bad Source, Good RAG"]["Bad Source + RAG"] += 1

        elif source_quality == "BAD" and rag_quality == "BAD":
            bad_source_bad_rag_count += 1

            if source_rag_quality == "GOOD":
                source_rag_combinations["Bad Source, Bad RAG"]["Good Source + RAG"] += 1

            elif source_rag_quality == "BAD":
                source_rag_combinations["Bad Source, Bad RAG"]["Bad Source + RAG"] += 1

    # In Row 8: Heading Labels
    # Col 3 (Bad Frequency), Col 5 (Good Frequency) for Source
    worksheet.cell(row=8, column=3, value="Bad Frequency")
    worksheet.cell(row=8, column=5, value="Good Frequency")
    worksheet.cell(row=9, column=3, value=bad_source_count)
    worksheet.cell(row=9, column=5, value=good_source_count)

    # Col 8 (Bad Frequency), Col 10 (Good Frequency) for RAG
    worksheet.cell(row=8, column=9, value="Bad Frequency")
    worksheet.cell(row=8, column=11, value="Good Frequency")
    worksheet.cell(row=9, column=9, value=bad_rag_count)
    worksheet.cell(row=9, column=11, value=good_rag_count)

    # Col 13 (Bad Frequency), Col 15 (Good Frequency) for Source + RAG
    worksheet.cell(row=8, column=15, value="Bad Frequency")
    worksheet.cell(row=8, column=17, value="Good Frequency")
    worksheet.cell(row=9, column=15, value=bad_source_rag_count)
    worksheet.cell(row=9, column=17, value=good_source_rag_count)


print("\nSummary Table:")
print(
    "{:<25} {:<20} {:<20} {:<20}".format(
        "Scenario", "Total Cases", "Good Source + RAG", "Bad Source + RAG"
    )
)
print("-" * 100)
print(
    "{:<25} {:<20} {:<20} {:<20}".format(
        "Good Source, Good RAG",
        good_source_good_rag_count,
        source_rag_combinations["Good Source, Good RAG"]["Good Source + RAG"],
        source_rag_combinations["Good Source, Good RAG"]["Bad Source + RAG"],
    )
)
print(
    "{:<25} {:<20} {:<20} {:<20}".format(
        "Good Source, Bad RAG",
        good_source_bad_rag_count,
        source_rag_combinations["Good Source, Bad RAG"]["Good Source + RAG"],
        source_rag_combinations["Good Source, Bad RAG"]["Bad Source + RAG"],
    )
)
print(
    "{:<25} {:<20} {:<20} {:<20}".format(
        "Bad Source, Good RAG",
        bad_source_good_rag_count,
        source_rag_combinations["Bad Source, Good RAG"]["Good Source + RAG"],
        source_rag_combinations["Bad Source, Good RAG"]["Bad Source + RAG"],
    )
)
print(
    "{:<25} {:<20} {:<20} {:<20}".format(
        "Bad Source, Bad RAG",
        bad_source_bad_rag_count,
        source_rag_combinations["Bad Source, Bad RAG"]["Good Source + RAG"],
        source_rag_combinations["Bad Source, Bad RAG"]["Bad Source + RAG"],
    )
)


# Save the modified workbook
workbook.save("sample_data/source_vs_rag1.xlsx")
